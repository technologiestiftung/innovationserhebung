
import json
import logging
import sys

from openpyxl import load_workbook
import pandas as pd

from config_importer import ConfigImporter
from parser import DataParserFactory

# TODO: Move filepaths to another file

sys.path.append("/app/sliders/config_importer")
logging.basicConfig(level=logging.INFO)

excel_file = "data/basisdaten_clean.xlsx"
outfile_path = "data/outfile.json"


class DataImporter:
    def import_data(self, data_file, config_file):
        """
        Main method of the class.
        Import the data from an Excel file.

        :param data_file: str, path to the Excel file containing the data to parse
        :param config_file: str, path to the YAML file containing the configuration
        """
        sheets = self.load_excel(data_file)
        config_importer = ConfigImporter()
        config = config_importer.get_config()

        output = {}
        data_parser_factory = DataParserFactory
        for plot in config:
            parser_type = config[plot]["parser_type"]
            parser = data_parser_factory.create_parser(parser_type)
            output[plot] = parser.parse(sheets, config)

        self.save_to_json(output, outfile_path)

    def load_excel(self, filepath):
        """
        Load Excel and convert to pandas DataFrames.

        :param filepath: str, path to the Excel file containing the data to parse
        :return: dict, where keys are names of sheets and values are pandas DataFrames
        """
        workbook = load_workbook(filename=filepath, data_only=True)
        sheets_data = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            data = sheet.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)

            # Remove rows and columns with empty values
            df = df.dropna(axis="index", how="all")
            df = df.dropna(axis="columns", how="all")

            sheets_data[sheet_name] = df

        return sheets_data

    def save_to_json(self, parsed_data, outfile_path):
        """
        Save data to JSON file.

        :param parsed_data: dict, parsed data
        :param outfile_path: str, path to the JSON file where to save the data
        """

        with open(outfile_path, "w") as outfile:
            json.dump(parsed_data, outfile)


# TODO: These two lines are for testing purposes. Remove later.
data_importer = DataImporter()
data_importer.import_data(excel_file, "config_file")
